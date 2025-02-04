import ssl
import datetime
import argparse
import time

from openpyxl.styles import Font
from openpyxl.workbook import Workbook
from pyvim import connect
from pyVmomi import vim
from modules.config_parser import int_host, int_user, int_pass
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

context = ssl._create_unverified_context()


def get_vcenter_connection(vcenter_ip, username, password):
    try:
        si = connect.SmartConnect(host=vcenter_ip, user=username, pwd=password, sslContext=context)
        return si
    except Exception as e:
        print(f"Failed to connect to vCenter: {e}")
        return None


def get_all_vms(content):
    container = content.viewManager.CreateContainerView(content.rootFolder, [vim.VirtualMachine], True)
    return container.view


def find_vm_by_ip(vms, target_ip):
    for vm in vms:
        try:
            if vm.guest.ipAddress == target_ip:
                return vm
        except Exception:
            pass
    return None


def get_vm_ip(vm):
    try:
        return vm.guest.ipAddress or "N/A"
    except Exception:
        return "N/A"


def extract_snapshot_info(snapshot_tree, snapshot_counter):

    snapshot_data = []

    for snapshot in snapshot_tree:
        snapshot_number = f"{snapshot_counter[0]})"
        snapshot_counter[0] += 1

        snapshot_name = f"{snapshot_number} {snapshot.name}"
        snapshot_desc = snapshot.description if snapshot.description else "No description added"


        snapshot_date = snapshot.createTime.replace(tzinfo=None)

        snapshot_data.append((snapshot_name, snapshot_desc, snapshot_date))


        if snapshot.childSnapshotList:
            snapshot_data.extend(extract_snapshot_info(snapshot.childSnapshotList, snapshot_counter))

    return snapshot_data


def wait_for_task_completion(task):
    """Wait until the vSphere task is complete."""
    print("Waiting for snapshot to complete...", end="", flush=True)

    while task.info.state not in [vim.TaskInfo.State.success, vim.TaskInfo.State.error]:
        time.sleep(2)  # Wait 2 seconds before checking again
        print(".", end="", flush=True)

    print("Done!")  # Print completion message

    if task.info.state == vim.TaskInfo.State.error:
        print(f"Snapshot failed: {task.info.error}")
        return False

    return True


def create_snapshot(vm, snapshot_name, snapshot_desc, memory_flag, quiesce_flag):
    """Create a snapshot for the given VM and wait for completion."""
    try:
        task = vm.CreateSnapshot_Task(
            name=snapshot_name,
            description=snapshot_desc,
            memory=memory_flag,  # Memory enabled by default
            quiesce=quiesce_flag
        )
        print(f"Creating snapshot '{snapshot_name}' for VM '{vm.name}' (IP: {vm.guest.ipAddress}) with Memory={memory_flag}...")

        # Wait for task to complete
        if wait_for_task_completion(task):
            print(f"Snapshot '{snapshot_name}' created successfully.")
            return True
        else:
            return False

    except Exception as e:
        print(f"Failed to create snapshot: {str(e)}")
        return False


parser = argparse.ArgumentParser(description="Create a VM snapshot and generate an Excel report.")
parser.add_argument("--ip", required=True, help="IP address of the VM.")
parser.add_argument("--name", required=True, help="Name for the snapshot.")
parser.add_argument("--no-memory", action="store_true", help="Disable VM memory in the snapshot.")
parser.add_argument("--quiesce", action="store_true", help="Quiesce the file system before snapshot.")
args = parser.parse_args()


vcenter_ip = int_host
username = int_user
password = int_pass

service_instance = get_vcenter_connection(vcenter_ip, username, password)
if not service_instance:
    exit(1)

content = service_instance.RetrieveContent()
vms = get_all_vms(content)

vm = find_vm_by_ip(vms, args.ip)
if vm:
    print(f"Found VM: {vm.name} (IP: {args.ip})")

    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    snapshot_name = args.name
    full_description = f"Snapshot_by_python_script--{timestamp}"

    memory_flag = not args.no_memory

    if not create_snapshot(vm, snapshot_name, full_description, memory_flag, args.quiesce):
        print("Snapshot creation failed, exiting.")
        exit(1)
else:
    print(f"No VM found with IP address: {args.ip}")
    exit(1)


excel_filename = f"vm_snapshots_{timestamp}.xlsx"
wb = Workbook()
ws = wb.active
ws.title = "VM Snapshots"

headers = ["VM Name", "VM IP", "Number of Snapshots", "Snapshot Name", "Description", "Created On"]
ws.append(headers)

bold_font = Font(bold=True)
for col in range(1, len(headers) + 1):
    ws.cell(row=1, column=col).font = bold_font

for vm in vms:
    vm_ip = get_vm_ip(vm)

    if vm.snapshot is not None:
        snapshot_counter = [1]
        snapshot_data = extract_snapshot_info(vm.snapshot.rootSnapshotList, snapshot_counter)

        snapshot_data.sort(key=lambda x: x[2], reverse=True)

        num_snapshots = len(snapshot_data)
        first_snapshot = snapshot_data[0]

        row = ws.max_row + 1
        ws.append([vm.name, vm_ip, num_snapshots, first_snapshot[0], first_snapshot[1], first_snapshot[2]])

        if vm_ip == args.ip:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).font = bold_font  # Apply bold formatting

        for snap in snapshot_data[1:]:
            ws.append(["", "", "", snap[0], snap[1], snap[2]])


for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal="left")

for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)

    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))

    ws.column_dimensions[col_letter].width = max_length + 2

# Save Excel file
wb.save(excel_filename)


print(f"Snapshot details saved to {excel_filename}")

connect.Disconnect(service_instance)
