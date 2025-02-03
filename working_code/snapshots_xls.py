import datetime
import ssl
from pyvim import connect
from pyVmomi import vim
from openpyxl import Workbook
from openpyxl.styles import Font
from modules.config_parser import int_host, int_user, int_pass

# Disable SSL verification for self-signed certificates
context = ssl._create_unverified_context()


def get_vcenter_connection(vcenter_ip, username, password):
    """Connect to vCenter and return the service instance."""
    si = connect.SmartConnect(host=vcenter_ip, user=username, pwd=password, sslContext=context)
    return si


def get_all_vms(content):
    """Retrieve all VMs from vCenter."""
    container = content.viewManager.CreateContainerView(content.rootFolder, [vim.VirtualMachine], True)
    return container.view


def get_vm_ip(vm):
    """Retrieve the IP address of the VM (if available)."""
    try:
        ip_address = vm.guest.ipAddress
        return ip_address if ip_address else None  # Return None if no IP
    except Exception:
        return None


def extract_snapshot_info(snapshot_tree, snapshot_counter):
    """Recursively extracts snapshot details (sequential numbering)."""
    snapshot_data = []

    for snapshot in snapshot_tree:
        snapshot_number = f"{snapshot_counter[0]})"
        snapshot_counter[0] += 1  # Increment global counter

        snapshot_name = f"{snapshot_number} {snapshot.name}"
        snapshot_desc = f"{snapshot_number} {snapshot.description if snapshot.description else 'No description added'}"
        snapshot_date = snapshot.createTime  # Keep as datetime object for sorting

        snapshot_data.append((snapshot_name, snapshot_desc, snapshot_date))

        # Recursively process child snapshots if they exist
        if snapshot.childSnapshotList:
            child_data = extract_snapshot_info(snapshot.childSnapshotList, snapshot_counter)
            snapshot_data.extend(child_data)

    return snapshot_data

vcenter_ip = int_host
username = int_user
password = int_pass

# Connect to vCenter
service_instance = get_vcenter_connection(vcenter_ip, username, password)
content = service_instance.RetrieveContent()

# Get VMs and process snapshots
vms = get_all_vms(service_instance.RetrieveContent())
timestamp = datetime.datetime.now().strftime("%Y_%m_%d_%H_%M_%S")
excel_filename = f"vm_snapshots_{timestamp}.xlsx"

# Create an Excel workbook and sheet
wb = Workbook()
ws = wb.active
ws.title = "VM Snapshots"

# Write headers and make them bold
headers = ["VM Name", "VM IP", "Number of Snapshots", "Snapshot Name", "Description", "Created On"]
ws.append(headers)

# Apply bold font to headers
bold_font = Font(bold=True)
for col in range(1, len(headers) + 1):
    ws.cell(row=1, column=col).font = bold_font  # Make header row bold

for vm in vms:
    vm_ip = get_vm_ip(vm)  # Get the IP address of the VM

    # Skip VMs with no IP (unknown)
    if vm_ip is None:
        continue

    if vm.snapshot is not None:
        # Initialize snapshot counter (using list so it updates in recursion)
        snapshot_counter = [1]

        # Extract all snapshots
        snapshot_data = extract_snapshot_info(vm.snapshot.rootSnapshotList, snapshot_counter)

        # Sort snapshots by "Created On" in descending order (latest first)
        snapshot_data.sort(key=lambda x: x[2], reverse=True)

        num_snapshots = len(snapshot_data)

        # Identify latest snapshot
        latest_snapshot = snapshot_data[0] if snapshot_data else None

        # Write the first snapshot entry in the VM row
        first_snapshot = snapshot_data[0]
        row = ws.max_row + 1
        ws.append([vm.name, vm_ip, num_snapshots, first_snapshot[0], first_snapshot[1], str(first_snapshot[2])])

        # Apply bold formatting to VM Name (Column A) & IP (Column B)
        ws[f"A{row}"].font = bold_font
        ws[f"B{row}"].font = bold_font

        # Apply bold formatting to the latest snapshot (Columns D, E, F)
        ws[f"D{row}"].font = bold_font
        ws[f"E{row}"].font = bold_font
        ws[f"F{row}"].font = bold_font

        # Write remaining snapshots for the VM in new rows (without repeating VM name/IP)
        for snap in snapshot_data[1:]:
            row = ws.max_row + 1
            ws.append(["", "", "", snap[0], snap[1], str(snap[2])])

# Save Excel file
wb.save(excel_filename)

print(f"Snapshot details saved to {excel_filename}")
